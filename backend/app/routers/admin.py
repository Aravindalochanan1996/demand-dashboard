from datetime import datetime, date
from io import BytesIO

from bson import ObjectId
from bson.errors import InvalidId
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from openpyxl import load_workbook

from ..database import clients_collection, rows_collection, users_collection
from ..auth import require_admin, hash_password
from ..schemas import RowOut, CreateUserRequest, UserOut, BulkDeleteRequest
from .clients import serialize_row

router = APIRouter(prefix="/api/admin", tags=["admin"])


def _oid(id_str: str) -> ObjectId:
    try:
        return ObjectId(id_str)
    except InvalidId:
        raise HTTPException(status_code=400, detail="Invalid id")


def _cell_to_value(value):
    """Excel cells can hold numbers, strings, dates, or None - normalize dates
    to plain strings so they're JSON-safe and editable like any other cell."""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


@router.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload a .xlsx file")

    content = await file.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)

    summary = []

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                summary.append({"client": sheet_name, "rows_imported": 0, "note": "empty sheet, skipped"})
                continue

            # Headers are taken as-is from row 1 - whatever text is there becomes
            # a column. Blank header cells are ignored (their column is dropped),
            # everything else is kept, in the order it appears.
            column_positions = []  # list of (source_index, header_text)
            for idx, h in enumerate(header_row):
                if h is not None and str(h).strip():
                    column_positions.append((idx, str(h).strip()))

            if not column_positions:
                summary.append({"client": sheet_name, "rows_imported": 0, "note": "no headers found, skipped"})
                continue

            columns = [c[1] for c in column_positions]
            first_col_idx = column_positions[0][0]

            parsed_rows = []
            for raw_row in rows_iter:
                if raw_row is None or all(v is None for v in raw_row):
                    continue
                first_val = raw_row[first_col_idx] if first_col_idx < len(raw_row) else None
                if first_val is not None and str(first_val).strip().lower() == "total":
                    continue

                row_data = {}
                for idx, header in column_positions:
                    value = raw_row[idx] if idx < len(raw_row) else None
                    row_data[header] = _cell_to_value(value)
                parsed_rows.append(row_data)

            # Find or create the client for this sheet
            client_doc = await clients_collection.find_one({"name": sheet_name})
            if not client_doc:
                result = await clients_collection.insert_one({"name": sheet_name, "columns": columns})
                client_id = result.inserted_id
            else:
                client_id = client_doc["_id"]
                await clients_collection.update_one(
                    {"_id": client_id}, {"$set": {"columns": columns}}
                )

            # This upload replaces the *active* rows for this client with the sheet's
            # contents. Rows already soft-deleted stay in the deleted-rows archive.
            await rows_collection.delete_many({"client_id": client_id, "deleted": {"$ne": True}})

            now = datetime.utcnow().isoformat()
            if parsed_rows:
                docs = [
                    {
                        "client_id": client_id,
                        "data": row_data,
                        "deleted": False,
                        "last_edited_by": admin["username"],
                        "last_edited_at": now,
                    }
                    for row_data in parsed_rows
                ]
                await rows_collection.insert_many(docs)

            summary.append({"client": sheet_name, "rows_imported": len(parsed_rows)})

        except Exception as exc:  # one bad sheet must never block the rest
            summary.append({"client": sheet_name, "error": str(exc)})
            continue

    return {"message": "Upload complete", "summary": summary}


@router.delete("/clients/{client_id}/rows/{row_id}", response_model=RowOut)
async def delete_row(client_id: str, row_id: str, admin: dict = Depends(require_admin)):
    row_doc = await rows_collection.find_one(
        {"_id": _oid(row_id), "client_id": _oid(client_id), "deleted": {"$ne": True}}
    )
    if not row_doc:
        raise HTTPException(status_code=404, detail="Row not found")

    await rows_collection.update_one(
        {"_id": row_doc["_id"]},
        {
            "$set": {
                "deleted": True,
                "deleted_by": admin["username"],
                "deleted_at": datetime.utcnow().isoformat(),
            }
        },
    )
    updated = await rows_collection.find_one({"_id": row_doc["_id"]})
    return serialize_row(updated)


@router.post("/clients/{client_id}/rows/bulk-delete", response_model=list[RowOut])
async def bulk_delete_rows(
    client_id: str, payload: BulkDeleteRequest, admin: dict = Depends(require_admin)
):
    ids = [_oid(rid) for rid in payload.row_ids]
    if not ids:
        raise HTTPException(status_code=400, detail="No row_ids provided")

    now = datetime.utcnow().isoformat()
    await rows_collection.update_many(
        {"_id": {"$in": ids}, "client_id": _oid(client_id), "deleted": {"$ne": True}},
        {"$set": {"deleted": True, "deleted_by": admin["username"], "deleted_at": now}},
    )
    cursor = rows_collection.find({"_id": {"$in": ids}, "client_id": _oid(client_id)})
    return [serialize_row(doc) async for doc in cursor]


@router.get("/clients/{client_id}/rows/deleted", response_model=list[RowOut])
async def get_deleted_rows(client_id: str, admin: dict = Depends(require_admin)):
    cursor = rows_collection.find({"client_id": _oid(client_id), "deleted": True})
    return [serialize_row(doc) async for doc in cursor]


@router.get("/last-edited")
async def last_edited(admin: dict = Depends(require_admin)):
    """Returns the most recent edit across all clients, plus per-client last edit."""
    cursor = rows_collection.find(
        {"deleted": {"$ne": True}, "last_edited_by": {"$ne": None}}
    ).sort("last_edited_at", -1)

    most_recent = None
    per_client = {}
    async for doc in cursor:
        client_doc = await clients_collection.find_one({"_id": doc["client_id"]})
        client_name = client_doc["name"] if client_doc else str(doc["client_id"])
        entry = {
            "client": client_name,
            "row_id": str(doc["_id"]),
            "edited_by": doc.get("last_edited_by"),
            "edited_at": doc.get("last_edited_at"),
        }
        if most_recent is None:
            most_recent = entry
        if client_name not in per_client:
            per_client[client_name] = entry

    return {"most_recent": most_recent, "per_client": list(per_client.values())}


@router.post("/users", response_model=UserOut)
async def create_user(payload: CreateUserRequest, admin: dict = Depends(require_admin)):
    if payload.role not in ("admin", "business_user"):
        raise HTTPException(status_code=400, detail="role must be 'admin' or 'business_user'")

    existing = await users_collection.find_one({"username": payload.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")

    await users_collection.insert_one(
        {
            "username": payload.username,
            "password_hash": hash_password(payload.password),
            "role": payload.role,
            "assigned_clients": payload.assigned_clients,
        }
    )
    return UserOut(username=payload.username, role=payload.role, assigned_clients=payload.assigned_clients)


@router.get("/users", response_model=list[UserOut])
async def list_users(admin: dict = Depends(require_admin)):
    cursor = users_collection.find({})
    return [
        UserOut(
            username=doc["username"],
            role=doc["role"],
            assigned_clients=doc.get("assigned_clients", []),
        )
        async for doc in cursor
    ]
